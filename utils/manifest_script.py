import os
import sys
import xml.sax
from openpyxl import load_workbook, Workbook
import argparse
from tqdm import tqdm

class UsePermissionHandler(xml.sax.ContentHandler):

    def __init__(self):
        self.permissions = []

    def startElement(self, tag, attributes):
        self.CurrentData = tag
        if tag == "uses-permission":
            self.permissions.append(attributes["android:name"])
    
    def endElement(self, tag):
        self.CurrentData = ""

    def characters(self, content):
        None

def get_permissions(parser, Handler, res_path):
    
    permissions = []
    files = os.listdir(res_path)
    t_files = tqdm(files)
    for file in t_files:
        first_path = os.path.join(res_path, file)
        if os.path.isdir(first_path):
            for f in os.listdir(first_path):
                second_path = os.path.join(first_path, f)
                if os.path.isfile(second_path) and f == "AndroidManifest.xml":
                    cur_permission = []
                    cur_permission.append(file)
                    parser.parse(second_path)
                    cur_permission += Handler.permissions
                    Handler.permissions = []
                    permissions.append(cur_permission)
    return permissions


def write_permissions(permissions, res_path):

    excel_path = os.path.join(res_path, "result.xlsx")
    if os.path.exists(excel_path):
        excel_file = load_workbook(excel_path)
        sheet = excel_file["Sheet1"]
    else:
        excel_file = Workbook()
        sheet = excel_file.create_sheet("Sheet1")

    #write into file by colum
    # for col, col_data in enumerate(permissions, start=1): 
    #     for row, value in enumerate(col_data, start=1):
    #         sheet.cell(row, col, value)
    #write into file by row
    for row in permissions:
        sheet.append(row)
    excel_file.save(excel_path)
    print("*********解析完成*********")

if __name__ == "__main__":
    
    parser = argparse.ArgumentParser()
    parser.add_argument("--apk_path", type=str, default=None)
    parser.add_argument("--res_path", type=str, default=None)
    args = parser.parse_args()
    if (args.apk_path == None or args.res_path == None or args.apk_path == "" or args.res_path == ""):
        print("Path cannot be empty!!!")
        sys.exit()
    parser = xml.sax.make_parser()
    parser.setFeature(xml.sax.handler.feature_namespaces, 0)
    Handler = UsePermissionHandler()
    parser.setContentHandler(Handler)
    permissions = get_permissions(parser, Handler, args.res_path)
    write_permissions(permissions, args.res_path)